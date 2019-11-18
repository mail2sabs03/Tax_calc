from PIL import Image
import pytesseract
import sys
from pdf2image import convert_from_path
import os
import sys
import xlwt
from xlwt import Workbook

# Path of the pdf
from InvoiceSummary import InvoiceSummary
from openpyxl import load_workbook
from datetime import datetime


print("execution started: ", datetime.now())

# PDF_file = "XYZInc.pdf"
# PDF_file = "ICE Data Invoice.pdf"
Tax_file: str = "AS_complete+.xlsx"

tax_dictionary = {}
wb = load_workbook(filename=Tax_file)
us_tax_sheet = wb['US_sales_tax_data']
us_max_row = us_tax_sheet.max_row

ca_tax_sheet = wb['CA_sales_tax_data']
ca_max_row = ca_tax_sheet.max_row


def initialize_tax_dictionary():
    # wb = load_workbook(filename=Tax_file)
    print(wb.sheetnames)
    # tax_sheet = wb['sales_tax_data']
    max_col = us_tax_sheet.max_column
    # max_row = tax_sheet.max_row
    print("max_col: ", max_col, " max_row: ", us_max_row)
    # print("A16", us_tax_sheet['A16'].value)

    i: int
    for i in range(2, us_max_row + 1):
        zip_code_cell = us_tax_sheet.cell(row=i, column=1)
        sales_tax_cell = us_tax_sheet.cell(row=i, column=5)
        # print("zip code: ", zip_code.value)
        # print("sales_tax: ", sales_tax.value)
        zip_code = zip_code_cell.value
        sales_tax = sales_tax_cell.value
        # zip_code = zip_code.strip()
        # sales_tax = sales_tax.strip()

        tax_dictionary[int(zip_code)] = sales_tax
        # print("zip code: ", tax_dictionary.get(zip_code))
        # print("sales_tax: ", tax_dictionary.get(sales_tax))

    for i in range(2, ca_max_row + 1):
        state_cell = ca_tax_sheet.cell(row=i, column=1)
        sales_tax_cell = ca_tax_sheet.cell(row=i, column=6)
        # print("zip code: ", zip_code.value)
        # print("sales_tax: ", sales_tax.value)
        state = state_cell.value
        sales_tax = sales_tax_cell.value
        # zip_code = zip_code.strip()
        # sales_tax = sales_tax.strip()

        tax_dictionary[state] = sales_tax
        # print("zip code: ", tax_dictionary.get(zip_code))
        # print("sales_tax: ", tax_dictionary.get(sales_tax))


def convert_pdf_to_text(pdf_file):
    """
    Part #1 : Converting PDF to images
    """
    # Store all the pages of the PDF in a variable
    pages = convert_from_path(pdf_file, 500)
    # Counter to store images of each page of PDF to image
    image_counter = 1
    # Iterate through all the pages stored above
    for page in pages:
        # Declaring filename for each page of PDF as JPG
        # For each page, filename will be:
        # PDF page 1 -> page_1.jpg
        # PDF page 2 -> page_2.jpg
        # PDF page 3 -> page_3.jpg
        # ....
        # PDF page n -> page_n.jpg
        filename = "page_" + str(image_counter) + ".jpg"

        # Save the image of the page in system
        page.save(filename, 'JPEG')

        # Increment the counter to update filename
        image_counter = image_counter + 1
    ''' 
    Part #2 - Recognizing text from the images using OCR 
    '''
    3
    # Variable to get count of total number of pages
    filelimit = image_counter - 1
    # Creating a text file to write the output
    outfile = "out_text.txt"
    # Open the file in append mode so that
    # All contents of all images are added to the same file
    f = open(outfile, "w")
    # Iterate from 1 to total number of pages
    for i in range(1, filelimit + 1):
        # Set filename to recognize text from
        # Again, these files will be:
        # page_1.jpg
        # page_2.jpg
        # ....
        # page_n.jpg
        filename = "page_" + str(i) + ".jpg"

        # Recognize the text as string in image using pytesserct
        text = str(((pytesseract.image_to_string(Image.open(filename)))))

        # The recognized text is stored in variable text
        # Any string processing may be applied on text
        # Here, basic formatting has been done:
        # In many PDFs, at line ending, if a word can't
        # be written fully, a 'hyphen' is added.
        # The rest of the word is written in the next line
        # Eg: This is a sample text this word here GeeksF-
        # orGeeks is half on first line, remaining on next.
        # To remove this, we replace every '-\n' to ''.
        text = text.replace('-\n', '')

        # Finally, write the processed text to the file.
        f.write(text)
    # Close the file after writing all the text.
    f.close()


def identify_template():
    with open("out_text.txt") as fp:
        cnt = 0
        check_template1 = False
        for line in fp:
            line = line.strip()
            if not line:
                continue
            if cnt > 50:
                return 0
            print("line {} contents {}".format(cnt, line))

            if check_template1:
                match_str = "Code Description Use Provider Loc ID Period Quantity Unit Price Amount"
                if match_str == line:
                    return 1
                else:
                    check_template1 = False
            if "License" in line:
                check_template1 = True

            if "ICE Data Pricing & Reference Data, LLC" == line:
                return 2

            cnt += 1
    fp.close()
    return 0


def get_sales_tax_from_state_city(state, city):
    i: int
    for i in range(2, us_max_row + 1):
        state_cell = us_tax_sheet.cell(row=i, column=2)
        city_cell = us_tax_sheet.cell(row=i, column=4)

        if state_cell.value.strip() == state and city_cell.value.strip() == city:
            sales_tax_cell = us_tax_sheet.cell(row=i, column=5)
            return sales_tax_cell.value

    for i in range(2, us_max_row + 1):
        state_cell = us_tax_sheet.cell(row=i, column=2)
        city_cell = us_tax_sheet.cell(row=i, column=4)

        if state_cell.value.strip() == state and city in city_cell.value.strip():
            sales_tax_cell = us_tax_sheet.cell(row=i, column=5)
            return sales_tax_cell.value


# Workbook is created
output_workbook = Workbook()

# add_sheet is used to create sheet.
sheet1 = output_workbook.add_sheet('Sheet 1')

sheet1.write(0, 0, 'From Company Name')
sheet1.write(0, 1, 'Account Number')
sheet1.write(0, 2, 'Invoice Number')
sheet1.write(0, 3, 'Invoice Date')
sheet1.write(0, 4, 'To Company Name')
sheet1.write(0, 5, 'City')
sheet1.write(0, 6, 'State')
sheet1.write(0, 7, 'Zip Code')
sheet1.write(0, 8, 'Country')
sheet1.write(0, 9, 'Subtotal')
sheet1.write(0, 10, 'Sales Tax')
sheet1.write(0, 11, 'Subtotal With Tax')
row_counter = 1

to_company_name = None
city = None
state = None
zip_code = None
country = None
sales_tax = None
from_company_name = None
invoice_date = None
invoice_number = None
account_number = None


def parse_based_on_template1():
    global row_counter
    global to_company_name
    global city
    global state
    global zip_code
    global country
    global sales_tax
    global from_company_name
    global invoice_date
    global invoice_number
    global account_number

    # summary_list = []
    # invoice_summary = None
    match_str = "Code Description Use Provider Loc ID Period Quantity Unit Price Amount"
    match_str_count = 0
    total_without_tax = 0
    total_with_tax = 0

    fetch_company_summary = False
    fetch_sub_total = False
    fetch_invoice_date = False
    fetch_invoice_number = True
    fetch_from_company_name = False
    fetch_account_number = False
    # wb.save('xlwt example.xls')

    with open("out_text.txt") as fp:
        for line in fp:
            line = line.strip()
            if not line:
                continue
            if fetch_invoice_number:
                if "Invoice Number:" in line:
                    invoice_number_summary = line.split('Invoice Number:')
                    invoice_number = invoice_number_summary[1].strip()
                    print("invoice number: ", invoice_number)
                    fetch_invoice_number = False
                    fetch_from_company_name = True
                continue
            if fetch_from_company_name:
                if "ATTN:" in line:
                    from_company_name = ''
                    from_company_name_summary = line.split(' ')
                    if len(from_company_name_summary) > 2:
                        for i in range(2, len(from_company_name_summary)):
                            from_company_name = from_company_name + from_company_name_summary[i] + " "
                    print("from company name: ", from_company_name)
                    fetch_from_company_name = False
                    fetch_account_number = True
                continue
            if fetch_account_number:
                if "Account Number" in line:
                    account_number_summary = line.split('Account Number')
                    print("account_number_summary: ", account_number_summary)
                    account_number_summary_data = account_number_summary[1].split(' ')
                    print("account_number_summary_data: ", account_number_summary_data)
                    account_number = account_number_summary_data[1].strip()
                    print("account number: ", account_number)
                    fetch_account_number = False
                    fetch_invoice_date = True
                    continue
            if fetch_invoice_date:
                if "Invoice Date" in line:
                    invoice_date_summary = line.split('Invoice Date')
                    print("invoice_date_summary: ", invoice_date_summary)
                    invoice_date_data = invoice_date_summary[1].split(' ')
                    print("invoice_date_data: ", invoice_date_data)
                    invoice_date = "" + invoice_date_data[1] + " " + invoice_date_data[2] + " " + invoice_date_data[3]
                    print("invoice date: ", invoice_date)
                    fetch_invoice_date = False
                    continue
            if fetch_company_summary:
                company_summary = line.split(',')
                if len(company_summary) < 5:
                    continue

                to_company_name = company_summary[0].strip()
                city = company_summary[-4].strip()
                state = company_summary[-3].strip()
                zip_code = company_summary[-2].strip()
                country = company_summary[-1].strip()

                print("Company Name: ", to_company_name)
                print("City: ", city)
                print("State: ", state)
                print("Zip Code: ", zip_code)
                print("Country: ", country)
                zip_code_arr = zip_code.split('-')
                zip_code = zip_code_arr[0]

                if country == 'US':
                    if zip_code and not zip_code.isspace() and zip_code.isnumeric():
                        sales_tax = tax_dictionary.get(int(zip_code))
                        print("sales tax: ", sales_tax)
                    else:
                        sales_tax = get_sales_tax_from_state_city(state, city)
                        print("sales tax: ", sales_tax)
                elif country == 'CA':
                    sales_tax = tax_dictionary.get(state)
                    print("sales tax: ", sales_tax)

                # block summary
                # invoice_summary = InvoiceSummary(to_company_name,
                #                                  city,
                #                                  state,
                #                                  zip_code,
                #                                  country,
                #                                  )
                # TODO: get tax based on the zip code and set it in the invoice summary
                fetch_company_summary = False
                fetch_sub_total = True
                continue

            if fetch_sub_total:
                if "Subtotal" in line:
                    sub_total_line = line.split()
                    subtotal = sub_total_line[-1]
                    print("Subtotal: ", subtotal)

                    subtotal_float = float(subtotal.replace(',', ''))
                    sales_tax_float = float(sales_tax)
                    sub_total_tax_float = subtotal_float * sales_tax_float
                    print("Subtotal tax: ", sub_total_tax_float)
                    print("Subtotal with tax: ", float(sub_total_line[-1].replace(',', '')) + sub_total_tax_float)

                    sheet1.write(row_counter, 0, from_company_name)
                    sheet1.write(row_counter, 1, account_number)
                    sheet1.write(row_counter, 2, invoice_number)
                    sheet1.write(row_counter, 3, invoice_date)
                    sheet1.write(row_counter, 4, to_company_name)
                    sheet1.write(row_counter, 5, city)
                    sheet1.write(row_counter, 6, state)
                    sheet1.write(row_counter, 7, zip_code)
                    sheet1.write(row_counter, 8, country)
                    sheet1.write(row_counter, 9, sub_total_line[-1])
                    sheet1.write(row_counter, 10, sales_tax_float)
                    sheet1.write(row_counter, 11, float(sub_total_line[-1].replace(',', '')) + sub_total_tax_float)
                    row_counter = row_counter + 1

                    total_without_tax = total_without_tax + float(sub_total_line[-1].replace(',', ''))
                    print("total without tax: ", total_without_tax)

                    total_with_tax = total_with_tax + float(sub_total_line[-1].replace(',', '')) + sub_total_tax_float
                    print("total with tax: ", total_with_tax)

                    print("---------------------------------------------------------------")
                    # invoice_summary.set_sub_total(sub_total_line[-1])
                    # TODO: compute subtotal with tax and set it to the invoice summary
                    # summary_list.append(invoice_summary)
                    fetch_sub_total = False
                    fetch_company_summary = True
                    continue

            if match_str == line and match_str_count == 0:
                fetch_company_summary = True
                match_str_count = 1
                continue

            # if "Please remit to:" in line:
            #    break
    fp.close()
    print("total without tax: ", total_without_tax)
    print("total with tax: ", total_with_tax)
    # sheet1.write(row_counter, 5, total_without_tax)
    # sheet1.write(row_counter, 7, total_with_tax)
    output_workbook.save('Tax Calculation.xls')


def parse_based_on_template2():
    global row_counter
    global to_company_name
    global city
    global state
    global zip_code
    global country
    global from_company_name
    global sales_tax
    global invoice_date
    global invoice_number
    global account_number

    fetch_company_summary = False
    fetch_address_line1 = False
    fetch_state_city_zip = False
    fetch_country = False
    fetch_subtotal = False
    with open("out_text.txt") as fp:
        for line in fp:
            line = line.strip()
            if not line:
                continue
            if "Invoice Date:" in line:
                invoice_date_summary = line.split('Invoice Date:')
                invoice_date = invoice_date_summary[1].strip()
                print("invoice date: ", invoice_date)
                fetch_company_summary = True
                continue
            if fetch_company_summary:
                company_summary = line.split('LLC')
                if len(company_summary) > 2:
                    from_company_name = company_summary[0].strip() + ' LLC'
                    print("from company name: ", from_company_name)
                    to_company_name = company_summary[1].strip() + ' LLC'
                    print("company name: ", to_company_name)
                    fetch_company_summary = False
                    fetch_address_line1 = True
                    continue
            if fetch_address_line1:
                if "Invoice No:" in line:
                    invoice_number_summary = line.split('Invoice No:')
                    invoice_number = invoice_number_summary[1].strip()
                    print("invoice number: ", invoice_number)
                fetch_address_line1 = False
                fetch_state_city_zip = True
                continue
            if fetch_state_city_zip:
                # Denver, CO 80206 NEW YORK, NY 10017
                add_summary = line.split(',')
                if len(add_summary) > 2:
                    element2 = add_summary[1].strip()
                    element2_arr = element2.split(' ')
                    print("element2:")
                    for i in range(0, len(element2_arr)):
                        print(element2_arr[i])
                    city = ''
                    for i in range(2, len(element2_arr)):
                        city = city + element2_arr[i] + " "
                    print("city: ", city)

                    state_zip = add_summary[2].strip()
                    state_zip_arr = state_zip.split(' ')
                    state = state_zip_arr[0].strip()
                    zip_code = state_zip_arr[1].strip()
                    print("state: ", state)
                    print("zipCode: ", zip_code)

                    sales_tax = tax_dictionary.get(int(zip_code))
                    print("sales tax: ", sales_tax)

                fetch_state_city_zip = False
                fetch_country = True
                continue
            if fetch_country:
                country_summary = line.split(' ')
                country = country_summary[1].strip()
                print("country: ", country)

                if "Account ID:" in line:
                    account_id_summary = line.split('Account ID:')
                    account_number = account_id_summary[1].strip()
                    print("invoice number: ", account_number)

                fetch_country = False
                fetch_subtotal = True
                continue
            if fetch_subtotal:
                if "Sub Total:" in line:
                    subtotal_summary = line.split(':')
                    subtotal = subtotal_summary[1].strip()
                    subtotal = subtotal.replace(',', '')
                    subtotal = subtotal.replace('$', '')
                    subtotal_float = float(subtotal)
                    print("subtotal: ", subtotal_float)

                    sales_tax_float = float(sales_tax)
                    sub_total_tax_float = subtotal_float * sales_tax_float
                    print("Subtotal tax: ", sub_total_tax_float)
                    print("Subtotal with tax: ", subtotal_float + sub_total_tax_float)

                    sheet1.write(row_counter, 0, from_company_name)
                    sheet1.write(row_counter, 1, account_number)
                    sheet1.write(row_counter, 2, invoice_number)
                    sheet1.write(row_counter, 3, invoice_date)
                    sheet1.write(row_counter, 4, to_company_name)
                    sheet1.write(row_counter, 5, city)
                    sheet1.write(row_counter, 6, state)
                    sheet1.write(row_counter, 7, zip_code)
                    sheet1.write(row_counter, 8, country)
                    sheet1.write(row_counter, 9, subtotal_float)
                    sheet1.write(row_counter, 10, sales_tax_float)
                    sheet1.write(row_counter, 11, subtotal_float + sub_total_tax_float)
                    row_counter = row_counter + 1

                    fetch_subtotal = False
                    break

        fp.close()
        output_workbook.save('Tax Calculation.xls')


initialize_tax_dictionary()
print("tax dictionary: ", tax_dictionary)

for filename in os.listdir("PDF_Files"):
    if filename.endswith(".pdf"):
        convert_pdf_to_text("PDF_Files/" + filename)
        template = identify_template()
        print(template)

        if template == 1:
            parse_based_on_template1()
        elif template == 2:
            parse_based_on_template2()

print("execution completed: ", datetime.now())
